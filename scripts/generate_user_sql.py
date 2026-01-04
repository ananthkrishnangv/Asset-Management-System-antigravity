#!/usr/bin/env python3
"""
Script to parse Excel employee list and generate SQL INSERT statements
"""
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# Load workbook
wb = openpyxl.load_workbook('/home/ananthakrishnan/Documents/Asset Management/Asset Management - Antigravity/Employees_Email_ID_CSIR_res_in_02_12_25.xlsx')
ws = wb.active

# Print header info
print("-- Employee Import SQL")
print("-- Generated from Employees_Email_ID_CSIR_res_in_02_12_25.xlsx")
print("-- Password for all users: Welcome@123")
print("")

# Get header row
headers = [cell.value for cell in ws[1]]
print(f"-- Columns found: {headers}")
print("")

# Default password hash for 'Welcome@123'
default_password_hash = "$2y$12$LN1c5/sVqxE6G8UqN0A1eO3M.qR7tHmCbKZwXjVpNlCaTnPfWZIZK"

inserts = []

for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row or not row[0]:  # Skip empty rows
        continue
    
    # Assuming columns: [0]=Name or AMS ID, [1]=Designation, [2]=Email or Name, etc.
    # Will adapt based on actual column headers
    # Let's try to map intelligently
    ams_id = None
    emp_name = None
    email_id = None
    designation = None
    
    for i, val in enumerate(row):
        if val is None:
            continue
        val_str = str(val).strip()
        if not val_str:
            continue
            
        header = headers[i].lower() if i < len(headers) and headers[i] else ""
        
        if "ams" in header or "id" in header or "emp" in header and "name" not in header:
            if val_str.isdigit() or (len(val_str) <= 10 and val_str.replace('-','').isdigit()):
                ams_id = val_str
        elif "name" in header:
            emp_name = val_str
        elif "email" in header or "@" in val_str:
            email_id = val_str
        elif "designation" in header or "desig" in header:
            designation = val_str
        else:
            # Fallback: first numeric looking value is AMS ID, first alpha is name
            if not ams_id and (val_str.isdigit() or val_str.replace('-','').isdigit()):
                ams_id = val_str
            elif not emp_name and val_str.isalpha() or ' ' in val_str:
                emp_name = val_str
    
    # Generate synthetic AMS ID if missing
    if not ams_id:
        ams_id = f"EMP{row_num:04d}"
    
    # Skip if no email
    if not email_id:
        continue
    
    # Use email prefix as name if no name
    if not emp_name:
        emp_name = email_id.split('@')[0].replace('.', ' ').replace('_', ' ').title()
    
    # Escape single quotes
    emp_name = emp_name.replace("'", "''")
    email_id = email_id.replace("'", "''")
    ams_id = ams_id.replace("'", "''")
    
    insert = f"('{ams_id}', '{emp_name}', '{email_id}', '{default_password_hash}', 'employee', NULL)"
    inserts.append(insert)

print("INSERT INTO `users` (`ams_id`, `emp_name`, `email_id`, `password`, `role`, `department_id`) VALUES")
print(",\n".join(inserts[:50]))  # First 50 for preview
print(";")
print("")
print(f"-- Total users: {len(inserts)}")
